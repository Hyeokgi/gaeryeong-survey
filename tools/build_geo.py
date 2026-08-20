# _source 의 접근성분석 엑셀에서 필지 좌표를 뽑아 kim.html / moon.html 에 GEO 로 심는다
# 다시 돌려도 안전하다 — 이미 있는 GEO 블록은 새 것으로 갈아끼운다.
import openpyxl, re, io, os, sys

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO, "_source")

ADDR = re.compile(r"개령면\s+(\S+리)\s+(\S+)$")


def load_coords():
    """(마을, 지번) -> (경도, 위도). 엑셀 두 개를 합친다."""
    geo, blank = {}, []

    def put(v, jb, lng, lat):
        try:
            geo[(v, jb)] = (float(lng), float(lat))
        except (TypeError, ValueError):
            blank.append((v, jb))

    # 7개 리 — village / jibun 열이 따로 있다
    ws = openpyxl.load_workbook(os.path.join(SRC, "개령면_접근성분석결과.xlsx"),
                                read_only=True, data_only=True)["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    H = {str(c).strip(): i for i, c in enumerate(rows[0])}
    for r in rows[1:]:
        put(str(r[H["village"]]).strip(), str(r[H["jibun"]]).strip(),
            r[H["경도"]], r[H["위도"]])

    # 동부리 — 뒤늦게 따로 돌린 것이라 열 구성이 다르다. 전체주소에서 뽑는다.
    ws = openpyxl.load_workbook(os.path.join(SRC, "동부리_접근성분석결과_v2.xlsx"),
                                read_only=True, data_only=True)["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    H = {str(c).strip(): i for i, c in enumerate(rows[0])}
    for r in rows[1:]:
        m = ADDR.search(str(r[H["전체주소"]]).strip())
        if m:
            put(m.group(1), m.group(2), r[H["경도"]], r[H["위도"]])

    return geo, blank


def build(fname, geo):
    path = os.path.join(REPO, fname)
    t = io.open(path, encoding="utf-8").read()

    m = re.search(r"const PARCELS = (\[.*?\]);\n", t, re.S)
    parcels = __import__("json").loads(m.group(1))

    hit, miss = [], []
    for p in parcels:
        c = geo.get((p["village"], p["jibun"]))
        if c:
            hit.append('"%s":[%.6f,%.6f]' % (p["id"], c[0], c[1]))
        else:
            miss.append(p)

    block = ("/* 필지 중심 좌표 [경도, 위도] · _source 접근성분석 엑셀에서 뽑았다.\n"
             "   tools/build_geo.py 로 다시 만든다. 좌표가 없는 필지는 키 자체가 없다. */\n"
             "const GEO = {%s};\n" % ",".join(hit))

    old = re.search(r"/\* 필지 중심 좌표.*?\nconst GEO = \{.*?\};\n", t, re.S)
    if old:
        t = t[:old.start()] + block + t[old.end():]
    else:
        t = t[:m.end()] + block + t[m.end():]

    io.open(path, "w", encoding="utf-8", newline="").write(t)
    print("[OK] %-10s %d/%d 필지에 좌표 (%.1f%%), %d KB 추가"
          % (fname, len(hit), len(parcels), len(hit) / len(parcels) * 100, len(block) // 1024))
    for p in miss:
        print("       좌표 없음: %s %s (%d일차 %02d번)" % (p["village"], p["jibun"], p["day"], p["seq"]))


if __name__ == "__main__":
    geo, blank = load_coords()
    print("좌표 %d건 확보, 엑셀에서 칸이 비어 있던 것 %d건" % (len(geo), len(blank)))
    for f in ("kim.html", "moon.html"):
        build(f, geo)
